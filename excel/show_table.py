from __future__ import annotations

import argparse
import logging
import configparser
from pathlib import Path

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

CONFIG_PATH = Path(__file__).parent / "config.ini"


def load_config(sheet_name: str | None = None) -> dict:
    if not CONFIG_PATH.exists():
        return {"header_row": 1}

    parser = configparser.ConfigParser()
    parser.read(CONFIG_PATH, encoding="utf-8")

    cfg = dict(parser["default"]) if parser.has_section("default") else {}
    if sheet_name and parser.has_section(sheet_name):
        cfg.update(parser[sheet_name])

    return {k: int(v) for k, v in cfg.items()}


def read_sheet(
    path: Path,
    sheet_name: str | None = None,
    header_row: int = 1,
    data_row: int | None = None,
) -> list[list[str]]:
    """header_row: 1-based. data_row: 1-based, defaults to header_row + 1."""
    if data_row is None:
        data_row = header_row + 1

    if not path.exists():
        raise SystemExit(f"file not found: {path}")
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name and sheet_name not in wb.sheetnames:
        wb.close()
        raise SystemExit(
            f"sheet '{sheet_name}' が見つかりません\n"
            f"  利用可能: {', '.join(wb.sheetnames)}"
        )
    ws = wb[sheet_name] if sheet_name else wb.active
    logger.info("sheet: %s", ws.title)

    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == header_row or i >= data_row:
            rows.append([str(c) if c is not None else "" for c in row])
    wb.close()
    return rows


def format_table(rows: list[list[str]]) -> str:
    if not rows:
        return "(empty)"

    col_count = max(len(r) for r in rows)
    for r in rows:
        r.extend([""] * (col_count - len(r)))

    widths = [max(len(r[i]) for r in rows) for i in range(col_count)]
    lines = []
    for j, row in enumerate(rows):
        line = " | ".join(cell.ljust(w) for cell, w in zip(row, widths))
        lines.append(line)
        if j == 0:
            lines.append("-+-".join("-" * w for w in widths))
    return "\n".join(lines)


def main() -> None:
    parser = argparse.ArgumentParser(description="Excel table viewer")
    parser.add_argument("file", type=Path, help="path to .xlsx")
    parser.add_argument("-s", "--sheet", default=None, help="sheet name")
    args = parser.parse_args()

    cfg = load_config(args.sheet)
    rows = read_sheet(
        args.file,
        args.sheet,
        cfg.get("header_row", 1),
        cfg.get("data_row"),
    )
    print(format_table(rows))


if __name__ == "__main__":
    main()