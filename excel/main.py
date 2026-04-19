from __future__ import annotations

import logging
from dataclasses import dataclass
from difflib import get_close_matches

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


class SheetNotFoundError(Exception):
    """"""

def _resolve_sheet(wb, expected: str) -> Worksheet:
    if expected in wb.sheetnames:
        return wb[expected]

    candidates = get_close_matches(expected, wb.sheetnames, n=3, cutoff=0.5)
    sheets_list = ", ".join(candidates)

    if candidates:
        suggestion = ", ".join(candidates)
        raise SheetNotFoundError(
            f"シート '{expected}' が見つかりません"
            f"\n typo の可能性: {suggestion}"
            f"\n -> Excel でシート名を '{expected}' に修正してください"
            f"\n (全シート: {sheets_list})"
        )
    raise SheetNotFoundError(
        f"シート '{expected} が見つかりません"
        f"\n 全シート: {sheets_list}"
        f"\n -> Excel に '{expected}' シートを作成してください"
    )


class JobEntry:
    no: str
    report_id: str | None
    has_filename: bool
    new_filename: str
    src_folder_name: str
    encode: str
    skip: str


def _log_jobs(entries: list[JobEntry]) -> None:
    header = (
        f"{'No':<6} {'20'} {'filename?':<10} "
        f"{'new_filename':<30} {'src_folder_name':<30} "
        f"{e.encode:<10} {e.skip:<6}"
    )
    lines = [header, "-" * len(header)]

    for i, e in enumerate(entries, start=1):
        id_str = e.report_id or "(なし)"
        fn_flag = "True" if e.has_filename else "False"
        lines.append(
            f"{i:<6} {id_str:<20} {fn_flag:<10} "
            f"{e.new_filename:<30} {e.src_folder_name:<30} "
            f"{e.encode:<10} {e.skip:<6}"
        )
    lines.append(f"合計: {len(entries)} 件")
    logger.info("\n%s", "\n".join(lines))


if __name__ == "__main__":
    from datetime import datetime as _dt

    from .utils import setup_logging

    setup_logging()

